import tempfile

import openpyxl
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import redirect, render, get_object_or_404
from django.template.loader import render_to_string, get_template
from django.utils.timezone import now
from django.views.decorators.http import require_POST
from openpyxl.styles import Font, PatternFill, Alignment

from accounts.models import PermissionInterface, User
from materiels.models import Materiel, CategorieMateriel
from pannes.forms import AffectationPanneForm, PanneForm, FicheDeReparationForm
from pannes.models import Panne, AffectationPanne, Notification


# Create your views here.
@login_required
def liste_pannes_et_techniciens_view(request):
    pannes = Panne.objects.all()
    techniciens = User.objects.filter(role__name='technicien')
    pannes_non_attribuees = Panne.objects.filter(affectation_panne__isnull=True)
    pannes_attribuees = AffectationPanne.objects.select_related('panne', 'technicien', 'panne__user')

    context={
        'pannes':pannes,
        'techniciens':techniciens,
        'pannes_non_attribuees':pannes_non_attribuees,
        'pannes_attribuees':pannes_attribuees
    }

    return render(request, 'pannes/liste_pannes.html', context)

User = get_user_model()
def affecter_pannes_view(request):
    if request.method == 'POST':
        panne_id = request.POST.get('panne')
        technicien_id = request.POST.get('technicien')

        try:
            panne = Panne.objects.get(id=panne_id)
            technicien = User.objects.get(id=technicien_id, role__name='technicien')
            chef = request.user

            # Vérifie que l'affectation n'existe pas déjà
            if AffectationPanne.objects.filter(panne=panne, technicien=technicien).exists():
                messages.warning(request, "Cette panne est déjà attribuée à ce technicien.")
            else:
                AffectationPanne.objects.create(
                    panne=panne,
                    technicien=technicien,
                    attribue_par=chef,
                    statut_reparation= 'non_traitee'
                )

                Notification.objects.create(
                    utilisateur=technicien,
                    panne=panne,
                    message=f"Panne assignée à {panne.user}",
                )
                messages.success(request, "La panne a bien été atribuée")
        except Panne.DoesNotExist:
            messages.error(request, "Panne introuvable.")
        except User.DoesNotExist:
            messages.error(request, "Technicien introuvable.")

        return redirect('pannes:liste_pannes')
    return redirect('pannes:liste_pannes')

@login_required
def mes_pannes_view(request):
    # Vérifie les permissions
    permission = PermissionInterface.objects.filter(
        role=request.user.role,
        interface__name="mes_pannes"
    ).first()

    #if not permission or not permission.peut_ajouter:
        #return HttpResponseForbidden("Vous n'avez pas la permission d'accéder à cette page.")
    
    if request.method == 'POST':
        form = PanneForm(request.POST)
        if form.is_valid():
            panne = form.save(commit=False)
            panne.user = request.user
            panne.save()
            return redirect('pannes:mes_pannes')  # rechargement de la page avec la nouvelle panne
    else:
        form = PanneForm()

    categories = CategorieMateriel.objects.all()
    pannes = Panne.objects.filter(user=request.user).prefetch_related(
        Prefetch(
            'affectation_panne',
            queryset=AffectationPanne.objects.order_by('-date_affectation')
        )
    )
    return render(request, 'pannes/mes_pannes.html', {'form': form, 'pannes': pannes, 'categories':categories})

@login_required
def mes_pannes_attribuees(request):
    technicien = request.user
    affectations = AffectationPanne.objects.filter(technicien=technicien)

    pannes_non_traitees = affectations.filter(statut_reparation='non_traitee')
    pannes_en_cours = affectations.filter(statut_reparation='en_cours')
    pannes_terminees = affectations.filter(statut_reparation='terminee')
    return render(request, 'pannes/mes_pannes_attribuees.html',{
        "pannes_non_traitees": pannes_non_traitees,
        "pannes_en_cours": pannes_en_cours,
        "pannes_terminees": pannes_terminees,
    })

@login_required
def changer_statut_affectation(request, affectation_id):
    affectation = get_object_or_404(AffectationPanne, id=affectation_id, technicien=request.user)
    nouveau_statut = request.POST.get('nouveau_statut')

    if nouveau_statut in ['en_cours', 'terminee']:
        if affectation.statut_reparation == 'non_traitee' and nouveau_statut == 'en_cours':
            affectation.statut_reparation = 'en_cours'
            affectation.date_intervention = now()
            affectation.save()
        elif affectation.statut_reparation == 'en_cours' and nouveau_statut == 'terminee':
            affectation.statut_reparation = 'terminee'
            affectation.date_reparation = now()
            affectation.save()
    return redirect('pannes:mes_pannes_attr')

@require_POST
def prendre_en_charge(request, panne_id):
    panne = get_object_or_404(Panne, id=panne_id, status='non traitee')
    panne.status = 'en cours'
    panne.save()
    messages.success(request, f'Panne #{panne.id} prise en charge avec succes.')
    return redirect('pannes:mes-pannes-attr')

def liste_pannes_view(request):
    return render(request, 'pannes/liste_pannes.html')

def exporter_pannes_excel_view(request):
    #creation du classeur et de la feuille
    global cell
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des pannes"

    #ajout des en-tetes
    headers = ["#", "Titre", "Description", "Priorité", "Statut", "Date"]
    header_font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="FF007BFF", end_color="FF007BFF", fill_type="solid")

    #positionnement des en-tetes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    pannes = Panne.objects.all().order_by('-date_signalement')

    #ajout des declarations de pannes
    for row_num, panne in enumerate(pannes,2):
        ws.cell(row=row_num, column=1, value=row_num-1)
        ws.cell(row=row_num, column=2, value=panne.titre)
        ws.cell(row=row_num, column=3, value=panne.description)
        ws.cell(row=row_num, column=4, value=panne.priority)
        ws.cell(row=row_num, column=5, value=panne.status)
        ws.cell(row=row_num, column=6, value=panne.date_signalement.strftime('%Y-%m-%d %H:%M'))

    #redimensionnement des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        if cell.value:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 8)
        ws.column_dimensions[column].width = adjusted_width

    # Réponse HTTP pour téléchargement
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="pannes.xlsx"'
    wb.save(response)
    return response


def charger_materiels(request):
    categorie_id = request.GET.get('categorie_materiel')
    materiels = Materiel.objects.filter(categorie_id=categorie_id).order_by('name')
    html = render_to_string('pannes/includes/option_materiels.html', {'materiels': materiels})
    return HttpResponse(html)

@login_required
def creer_fiche_de_reparation(request, affectation_id):
    affect = get_object_or_404(AffectationPanne, id=affectation_id, technicien=request.user)

    initial_data = {
        'titre': affect.panne.titre,
        'description_panne': affect.panne.description,
        'date_affectation': affect.date_affectation,
        'date_intervention': affect.date_intervention,
        'date_reparation': affect.date_intervention if affect.statut_reparation == 'terminee' else '',
        'technicien': affect.technicien.username,
        'attribue_par': affect.attribue_par.username,
    }

    if request.method == 'POST':
        form = FicheDeReparationForm(request.POST, initial=initial_data)
        if form.is_valid():
            intervention = form.cleaned_data.get('description_intervention')
            affect.commentaire_intervention = intervention
            affect.save()
            return redirect('pannes:fiche_reparation_print', affectation_id=affect.id)
    else:
        # Ici on initialise form pour les requêtes GET
        form = FicheDeReparationForm(initial=initial_data)

    # Cette ligne sera toujours atteinte (POST invalide ou GET)
    return render(request, 'pannes/fiche_reparation_form.html', {
        'form': form,
        'affectation': affect,
    })

@login_required
def fiche_reparation_printable(request, affectation_id):
    affectation = get_object_or_404(AffectationPanne, id=affectation_id)
    today = now().date()
    return render(request, 'pannes/fiche_reparation_pdf.html', {
        'affectation': affectation,
        'today' : today,
    })


@login_required
def liste_fiches_reparation(request):
    if request.user.role.name == 'technicien':
        fiches = AffectationPanne.objects.filter(
            technicien=request.user,
            statut_reparation='terminee',
            commentaire_intervention__isnull=False
        )
    else:
        fiches = AffectationPanne.objects.filter(
            statut_reparation='terminee',
            commentaire_intervention__isnull=False
        )

    return render(request, 'pannes/mes_fiches_de_reparation.html', {'fiches': fiches})